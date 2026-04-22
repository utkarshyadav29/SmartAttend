from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify
from flask_login import login_required, current_user
from functools import wraps
from extensions import db
from models import User, Department, Class, Subject, Student, AttendanceRecord, ApprovalRequest
from datetime import datetime, date, timedelta
import csv, io, json

admin_bp = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated

@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    total_students = Student.query.count()
    total_classes = Class.query.count()
    total_teachers = User.query.filter_by(role='teacher').count()
    pending_approvals = ApprovalRequest.query.filter_by(status='pending').count()
    today = date.today()
    today_records = AttendanceRecord.query.filter_by(date=today).count()

    # Weekly attendance data for chart
    weekly = []
    for i in range(7):
        d = today - timedelta(days=6-i)
        present = AttendanceRecord.query.filter_by(date=d, status='present').count()
        absent = AttendanceRecord.query.filter_by(date=d, status='absent').count()
        weekly.append({'date': d.strftime('%a'), 'present': present, 'absent': absent})

    recent_approvals = ApprovalRequest.query.filter_by(status='pending').order_by(ApprovalRequest.created_at.desc()).limit(5).all()

    # Monthly data from real DB
    months_data = []
    for i in range(6, 0, -1):
        month_start = date.today().replace(day=1) - timedelta(days=30 * (i-1))
        month_label = month_start.strftime('%b')
        total_m = AttendanceRecord.query.filter(db.func.strftime('%Y-%m', AttendanceRecord.date) == month_start.strftime('%Y-%m')).count()
        present_m = AttendanceRecord.query.filter(db.func.strftime('%Y-%m', AttendanceRecord.date) == month_start.strftime('%Y-%m'), AttendanceRecord.status=='present').count()
        avg = round(present_m / total_m * 100, 1) if total_m else 0
        months_data.append({'month': month_label, 'avg': avg})

    # Real faculty benchmark data from DB
    faculty_stats = []
    teachers = User.query.filter_by(role='teacher', is_active_account=True).all()
    for t in teachers:
        subj_ids = [s.id for s in t.subjects]
        if not subj_ids:
            continue
        for s in t.subjects:
            total = AttendanceRecord.query.filter_by(subject_id=s.id).count()
            present = AttendanceRecord.query.filter_by(subject_id=s.id, status='present').count()
            score = round(present / total * 100, 1) if total else 0
            faculty_stats.append({'name': t.name, 'subject': s.name.upper(), 'score': score})
        if len(faculty_stats) >= 4:
            break

    return render_template('admin/dashboard.html',
        total_students=total_students,
        total_classes=total_classes,
        total_teachers=total_teachers,
        pending_approvals=pending_approvals,
        today_records=today_records,
        weekly=json.dumps(weekly),
        monthly=json.dumps(months_data),
        recent_approvals=recent_approvals,
        faculty_stats=faculty_stats,
        institute_name="GH Raisoni College of Engineering & Management"
    )

@admin_bp.route('/classes', methods=['GET', 'POST'])
@login_required
@admin_required
def classes():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_dept':
            dept = Department(name=request.form['dept_name'], code=request.form['dept_code'].upper())
            db.session.add(dept)
            db.session.commit()
            flash('Department added successfully', 'success')
        elif action == 'add_class':
            cls = Class(name=request.form['class_name'], section=request.form.get('section',''),
                        year=int(request.form.get('year', 1)), department_id=int(request.form['dept_id']))
            db.session.add(cls)
            db.session.commit()
            flash('Class added successfully', 'success')
        elif action == 'add_subject':
            subj = Subject(name=request.form['subj_name'], code=request.form.get('subj_code',''),
                           class_id=int(request.form['class_id']), credits=int(request.form.get('credits',4)))
            db.session.add(subj)
            db.session.commit()
            flash('Subject added successfully', 'success')
        elif action == 'assign_teacher':
            subj = Subject.query.get(int(request.form['subject_id']))
            subj.teacher_id = int(request.form['teacher_id']) if request.form['teacher_id'] else None
            db.session.commit()
            flash('Teacher assigned successfully', 'success')
        return redirect(url_for('admin.classes'))

    departments = Department.query.all()
    teachers = User.query.filter_by(role='teacher').all()
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/delete_department/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_department(id):
    dept = Department.query.get_or_404(id)
    # Recursively handle classes and students if needed, or just delete the dept
    # To prevent foreign key errors, we might need to handle dependencies
    db.session.delete(dept)
    db.session.commit()
    flash(f'Department "{dept.name}" deleted.', 'success')
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/add_class', methods=['POST'])
@login_required
@admin_required
def add_class():
    dept_id = request.form.get('dept_id')
    section = request.form.get('section', '').strip()
    dept = Department.query.get(dept_id)
    if not dept or not section:
        flash('Invalid department or division name.', 'error')
        return redirect(url_for('admin.staff_log'))
    
    # User requested: "divison will also contain the same name of the department only the divion will be changed"
    class_name = f"{dept.name}"
    cls = Class(name=class_name, section=section, department_id=dept.id, year=1)
    db.session.add(cls)
    db.session.commit()
    flash(f'Division "{section}" added to {dept.name}.', 'success')
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/delete_class/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_class(id):
    cls = Class.query.get_or_404(id)
    db.session.delete(cls)
    db.session.commit()
    flash(f'Division "{cls.section}" deleted.', 'success')
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/remove_faculty/<int:id>', methods=['POST'])
@login_required
@admin_required
def remove_faculty(id):
    user = User.query.get_or_404(id)
    # "authority of faculty will be taken... whole details... remain but it will not be a faculty"
    user.role = 'guest' 
    user.is_active_account = False
    db.session.commit()
    flash(f'Faculty status removed for {user.name}. Data preserved.', 'success')
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/settings')
@login_required
@admin_required
def settings():
    return render_template('admin/settings.html')

@admin_bp.route('/approvals', methods=['GET', 'POST'])
@login_required
@admin_required
def approvals():
    if request.method == 'POST':
        action = request.form.get('action')
        req_type = request.form.get('req_type')

        if req_type == 'teacher':
            teacher_id = int(request.form.get('teacher_id'))
            teacher = User.query.get(teacher_id)
            if teacher:
                if action == 'approve':
                    teacher.is_active_account = True
                    db.session.commit()
                    flash(f"Teacher {teacher.name}'s account approved.", "success")
                elif action == 'reject':
                    # Instead of deleting, we can just set a status or role
                    teacher.role = 'rejected'
                    db.session.commit()
                    flash(f"Teacher {teacher.name}'s account rejected.", "success")
        else:
            # Subject Approval Logic
            req_id = int(request.form.get('request_id'))
            req = ApprovalRequest.query.get(req_id)
            if req:
                req.status = 'approved' if action == 'approve' else 'rejected'
                req.reviewed_by = current_user.id
                req.reviewed_at = datetime.utcnow()
                if action == 'approve':
                    req.subject.teacher_id = req.teacher_id
                db.session.commit()
                flash(f'Request {req.status} successfully', 'success')
                
        return redirect(url_for('admin.approvals'))

    pending_teachers = User.query.filter_by(role='teacher', is_active_account=False).all()
    history_teachers = User.query.filter(User.role.in_(['teacher', 'rejected']), (User.is_active_account == True) | (User.role == 'rejected')).all()
    pending = ApprovalRequest.query.filter_by(status='pending').order_by(ApprovalRequest.created_at.desc()).all()
    history = ApprovalRequest.query.filter(ApprovalRequest.status != 'pending').order_by(ApprovalRequest.created_at.desc()).limit(20).all()
    return render_template('admin/approvals.html', pending=pending, history=history, pending_teachers=pending_teachers, history_teachers=history_teachers)


@admin_bp.route('/departments', methods=['POST'])
@login_required
@admin_required
def add_department():
    name = request.form.get('dept_name', '').strip()
    code = request.form.get('dept_code', '').strip().upper()
    if not name or not code:
        flash('Department name and code are required.', 'error')
        return redirect(url_for('admin.staff_log'))
    if Department.query.filter_by(name=name).first():
        flash(f'Department "{name}" already exists.', 'warning')
        return redirect(url_for('admin.staff_log'))
    dept = Department(name=name, code=code)
    db.session.add(dept)
    db.session.commit()
    flash(f'Department "{name}" created successfully.', 'success')
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/students', methods=['GET', 'POST'])
@login_required
@admin_required
def students():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add_student':
            st = Student(student_id=request.form['student_id'], name=request.form['name'],
                         class_id=int(request.form['class_id']), email=request.form.get('email',''))
            db.session.add(st)
            db.session.commit()
            flash('Student added', 'success')
        elif action == 'import_csv':
            file = request.files.get('csv_file')
            if file:
                stream = io.StringIO(file.stream.read().decode('utf-8'))
                reader = csv.DictReader(stream)
                count = 0
                for row in reader:
                    if not Student.query.filter_by(student_id=row.get('student_id','')).first():
                        st = Student(student_id=row['student_id'], name=row['name'],
                                     class_id=int(row.get('class_id', request.form.get('import_class_id',1))))
                        db.session.add(st)
                        count += 1
                db.session.commit()
                flash(f'Imported {count} students', 'success')
        return redirect(url_for('admin.students'))

    class_id = request.args.get('class_id', type=int)
    classes = Class.query.all()
    students = Student.query.filter_by(class_id=class_id).all() if class_id else Student.query.all()
    return redirect(url_for('admin.staff_log'))

@admin_bp.route('/import_students', methods=['POST'])
@login_required
@admin_required
def import_students():
    file = request.files.get('file')
    dept_name = request.form.get('dept_name', '')

    if not file or file.filename == '':
        flash('No file selected. Please choose a CSV or Excel file.', 'error')
        return redirect(url_for('admin.staff_log'))

    filename = file.filename.lower()
    
    try:
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            except ImportError:
                flash('openpyxl not installed. Please upload a CSV file instead.', 'error')
                return redirect(url_for('admin.staff_log'))
        else:
            flash('Invalid file type. Please upload a .csv or .xlsx file.', 'error')
            return redirect(url_for('admin.staff_log'))

        # Find or get a default class to assign students
        dept = Department.query.filter_by(name=dept_name).first()
        cls = None
        if dept:
            cls = Class.query.filter_by(department_id=dept.id).first()
        if not cls:
            cls = Class.query.first()
        if not cls:
            flash('No class found to assign students. Please ensure a class exists first.', 'error')
            return redirect(url_for('admin.staff_log'))

        count = 0
        for row in rows:
            sid = row.get('student_id', '').strip()
            name = row.get('name', '').strip()
            email = row.get('email', '').strip()
            if not sid or not name:
                continue
            if not Student.query.filter_by(student_id=sid).first():
                s = Student(student_id=sid, name=name, email=email, class_id=cls.id)
                db.session.add(s)
                count += 1

        db.session.commit()
        flash(f'Successfully imported {count} students.', 'success')

    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')

    return redirect(url_for('admin.staff_log'))


@admin_bp.route('/upload_photo/<int:student_id>', methods=['POST'])
@login_required
@admin_required
def upload_photo(student_id):
    from ai.detector import detect_and_encode_faces
    import os
    student = Student.query.get_or_404(student_id)
    files = request.files.getlist('photos')
    all_encodings = student.get_encoding()

    for f in files:
        if f and f.filename:
            path = os.path.join('/tmp', f'student_{student_id}_{f.filename}')
            f.save(path)
            encs = detect_and_encode_faces(path)
            all_encodings.extend(encs)
            os.remove(path)

    student.set_encoding(all_encodings)
    student.photo_count = len(files)
    db.session.commit()
    return jsonify({'success': True, 'count': len(all_encodings)})

@admin_bp.route('/analytics')
@login_required
@admin_required
def analytics():
    from sqlalchemy import func
    classes = Class.query.all()
    class_id = request.args.get('class_id', type=int)

    # Overall stats
    total_records = AttendanceRecord.query.count()
    present_count = AttendanceRecord.query.filter_by(status='present').count()
    overall_pct = round(present_count / total_records * 100, 1) if total_records else 0

    # Per-class breakdown
    class_stats = []
    for cls in classes:
        student_ids = [s.id for s in cls.students]
        if not student_ids:
            continue
        total = AttendanceRecord.query.filter(AttendanceRecord.student_id.in_(student_ids)).count()
        present = AttendanceRecord.query.filter(AttendanceRecord.student_id.in_(student_ids), AttendanceRecord.status=='present').count()
        
        subjects_data = []
        colors = ["var(--brand)", "var(--accent)", "#D6C3EB", "#2DA84F", "#FF9800", "#9C27B0"]
        c_idx = 0
        for subj in cls.subjects:
            s_tot = AttendanceRecord.query.filter(AttendanceRecord.subject_id==subj.id, AttendanceRecord.student_id.in_(student_ids)).count()
            s_pres = AttendanceRecord.query.filter(AttendanceRecord.subject_id==subj.id, AttendanceRecord.student_id.in_(student_ids), AttendanceRecord.status=='present').count()
            if s_tot > 0:
                s_pct = round((s_pres/s_tot)*100, 1)
                subjects_data.append({'name': subj.name, 'val': s_pct, 'color': colors[c_idx % len(colors)], 'weight': f"{s_tot} sessions"})
                c_idx += 1

        overall_cls_pct = round(present/total*100,1) if total else 0
        class_stats.append({
            'name': cls.full_name, 'total': total, 'present': present,
            'pct': overall_cls_pct,
            'overall': f"{overall_cls_pct}%",
            'subjects': subjects_data
        })

    # Faculty stats
    faculty_stats = []
    teachers = User.query.filter_by(role='teacher', is_active_account=True).all()
    for t in teachers:
        subj_ids = [s.id for s in t.subjects]
        if not subj_ids:
            continue
        tot_m = AttendanceRecord.query.filter(AttendanceRecord.subject_id.in_(subj_ids)).count()
        if tot_m > 0:
            pres_m = AttendanceRecord.query.filter(AttendanceRecord.subject_id.in_(subj_ids), AttendanceRecord.status=='present').count()
            score = round(pres_m / tot_m * 100, 1)
            faculty_stats.append({
                'name': t.name,
                'initials': t.name[:2].upper() if t.name else 'T',
                'dept': t.department or 'General',
                'total_sessions': tot_m,
                'score': score
            })

    # Sort faculty by score descending
    faculty_stats.sort(key=lambda x: x['score'], reverse=True)

    return render_template('admin/analytics.html', class_stats=json.dumps(class_stats),
                           overall_pct=overall_pct, total_records=total_records, present_count=present_count,
                           classes=classes, faculty_stats=faculty_stats)

@admin_bp.route('/staff_log', methods=['GET', 'POST'])
@login_required
@admin_required
def staff_log():
    if request.method == 'POST':
        t = User(username=request.form['username'], name=request.form['name'], role='teacher',
                 email=request.form.get('email',''), department=request.form.get('department',''),
                 employee_id=request.form.get('employee_id',''))
        t.set_password(request.form['password'])
        db.session.add(t)
        db.session.commit()
        flash('Teacher account created', 'success')
        return redirect(url_for('admin.staff_log'))
    teachers = User.query.filter_by(role='teacher').all()
    departments = Department.query.all()
    
    dept_stats = []
    for d in departments:
        cls_count = Class.query.filter_by(department_id=d.id).count()
        stu_count = Student.query.join(Class).filter(Class.department_id==d.id).count()
        dept_stats.append({
            'id': d.id,
            'code': d.code,
            'name': d.name,
            'classes': cls_count,
            'students': stu_count
        })

    return render_template('admin/staff_log.html', teachers=teachers, dept_stats=dept_stats)

@admin_bp.route('/get_divisions/<int:dept_id>')
@login_required
@admin_required
def get_divisions(dept_id):
    classes = Class.query.filter_by(department_id=dept_id).all()
    return jsonify([{'id': c.id, 'section': c.section} for c in classes])
